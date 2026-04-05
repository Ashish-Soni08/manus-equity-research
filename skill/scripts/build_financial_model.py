#!/usr/bin/env python3
"""
Build an equity research financial model Excel workbook for a given company.

Usage:
    python3 build_financial_model.py <TICKER> <OUTPUT_PATH> <DATA_JSON_PATH>

This script reads a structured JSON file containing the company's financial data
and produces a multi-sheet Excel workbook with:
  1. Income Statement (Historical + Projections)
  2. Balance Sheet Summary
  3. DCF Valuation Model
  4. Comparable Company Analysis
  5. Key Metrics Dashboard

The DATA_JSON_PATH should point to a JSON file with this structure:
{
  "ticker": "PLTR",
  "company_name": "Palantir Technologies Inc.",
  "years": [2020, 2021, ...],
  "revenue": [1093, 1542, ...],
  "cogs": [427, 600, ...],
  "sga": [...], "rd": [...], "ga": [...],
  "sbc": [...], "net_income": [...], "eps_diluted": [...],
  "projection_years": ["2026E", "2027E", ...],
  "proj_revenue": [...], "proj_cogs": [...], ...
  "balance_sheet": { ... },
  "dcf_assumptions": {
    "wacc": 0.134, "terminal_growth": 0.04, "beta": 1.65,
    "risk_free_rate": 0.043, "erp": 0.055,
    "shares_outstanding": 2573, "net_cash": 7177,
    "proj_fcf": [...]
  },
  "comps": [
    {"name": "CrowdStrike", "ticker": "CRWD", "ev_revenue": 22.7, ...},
    ...
  ]
}

NOTE: This is a template. Manus should populate the JSON data file during the
research phase, then run this script to generate the Excel workbook.
Adjust projections, assumptions, and comps based on the specific company.
"""

import sys
import json
import os

def main():
    if len(sys.argv) < 4:
        print("Usage: python3 build_financial_model.py <TICKER> <OUTPUT_PATH> <DATA_JSON_PATH>")
        print("\nThis script requires a structured JSON data file. See script docstring for format.")
        sys.exit(1)

    ticker = sys.argv[1].upper()
    output_path = sys.argv[2]
    data_path = sys.argv[3]

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    with open(data_path) as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # --- Style definitions ---
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    section_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    normal_font = Font(name='Calibri', size=10)
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thick_border = Border(bottom=Side(style='medium', color='1F4E79'))

    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Sheet 1: Income Statement ---
    ws1 = wb.active
    ws1.title = "Income Statement"
    ws1.merge_cells('A1:L1')
    ws1['A1'] = f'{data.get("company_name", ticker)} ({ticker}) — INCOME STATEMENT'
    ws1['A1'].font = title_font
    ws1['A2'] = 'All figures in USD millions unless otherwise stated'
    ws1['A2'].font = Font(name='Calibri', size=9, italic=True, color='666666')

    all_years = data.get('years', []) + data.get('projection_years', [])
    revenue = data.get('revenue', []) + data.get('proj_revenue', [])
    cogs = data.get('cogs', []) + data.get('proj_cogs', [])

    row = 4
    ws1.cell(row=row, column=1, value='($ in millions)')
    for i, y in enumerate(all_years):
        ws1.cell(row=row, column=i+2, value=y)
    style_header_row(ws1, row, len(all_years)+1)

    # Revenue row
    r = 5
    ws1.cell(row=r, column=1, value='Revenue').font = section_font
    for i, v in enumerate(revenue):
        cell = ws1.cell(row=r, column=i+2, value=v)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')

    ws1.column_dimensions['A'].width = 28
    for i in range(2, len(all_years)+2):
        ws1.column_dimensions[get_column_letter(i)].width = 14

    # --- Sheet 2: DCF Valuation ---
    ws3 = wb.create_sheet("DCF Valuation")
    ws3.merge_cells('A1:H1')
    ws3['A1'] = f'{ticker} — DCF VALUATION MODEL'
    ws3['A1'].font = title_font

    dcf = data.get('dcf_assumptions', {})
    assumptions = [
        ('WACC', dcf.get('wacc', 0.13)),
        ('Terminal Growth Rate', dcf.get('terminal_growth', 0.04)),
        ('Beta', dcf.get('beta', 1.5)),
        ('Risk-Free Rate', dcf.get('risk_free_rate', 0.043)),
        ('Equity Risk Premium', dcf.get('erp', 0.055)),
        ('Shares Outstanding (M)', dcf.get('shares_outstanding', 0)),
        ('Net Cash (M)', dcf.get('net_cash', 0)),
    ]

    r = 3
    ws3.cell(row=r, column=1, value='KEY ASSUMPTIONS').font = section_font
    style_header_row(ws3, r, 3)
    for label, val in assumptions:
        r += 1
        ws3.cell(row=r, column=1, value=label).font = normal_font
        cell = ws3.cell(row=r, column=2, value=val)
        cell.fill = input_fill
        if isinstance(val, float) and val < 1:
            cell.number_format = '0.0%'
        else:
            cell.number_format = '#,##0'

    # DCF calculation
    proj_fcf = dcf.get('proj_fcf', [])
    wacc = dcf.get('wacc', 0.13)
    tgr = dcf.get('terminal_growth', 0.04)
    shares = dcf.get('shares_outstanding', 1)
    net_cash = dcf.get('net_cash', 0)

    if proj_fcf and wacc > tgr:
        dfs = [1/(1+wacc)**i for i in range(1, len(proj_fcf)+1)]
        pv_fcf = sum(f*d for f, d in zip(proj_fcf, dfs))
        tv = proj_fcf[-1] * (1+tgr) / (wacc - tgr)
        pv_tv = tv * dfs[-1]
        ev = pv_fcf + pv_tv
        equity_val = ev + net_cash
        price_per_share = equity_val / shares

        r += 2
        ws3.cell(row=r, column=1, value='VALUATION OUTPUT').font = section_font
        style_header_row(ws3, r, 3)
        outputs = [
            ('PV of Projected FCF', pv_fcf),
            ('Terminal Value', tv),
            ('PV of Terminal Value', pv_tv),
            ('Enterprise Value', ev),
            ('+ Net Cash', net_cash),
            ('Equity Value', equity_val),
            ('Shares Outstanding (M)', shares),
            ('Implied Price per Share', price_per_share),
        ]
        for label, val in outputs:
            r += 1
            ws3.cell(row=r, column=1, value=label).font = normal_font
            cell = ws3.cell(row=r, column=2, value=val)
            if label == 'Implied Price per Share':
                cell.number_format = '$#,##0.00'
            elif label == 'Shares Outstanding (M)':
                cell.number_format = '#,##0'
            else:
                cell.number_format = '#,##0'

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 18

    # --- Save ---
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Financial model saved to {output_path}")

if __name__ == '__main__':
    main()
